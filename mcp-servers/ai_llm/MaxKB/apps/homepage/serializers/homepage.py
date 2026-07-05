# coding=utf-8
"""
    @project: MaxKB
    @Author：虎虎
    @file： homepage.py
    @date：2026/5/13 14:34
    @desc:
"""
import datetime
import os
from typing import List, Dict

import openpyxl
from django.db import models
from django.db.models import QuerySet, Count, Q, UUIDField, Sum, F, BigIntegerField, Value, ExpressionWrapper, \
    IntegerField, Window
from django.db.models.functions import Cast, Coalesce, RowNumber
from django.forms import CharField
from django.http import HttpResponse
from django.utils import timezone
from django.utils.translation import gettext_lazy as _, gettext
from rest_framework import serializers

from application.models import Application, ApplicationChatUserStats, Chat, ChatRecord
from common.constants.permission_constants import RoleConstants
from common.db.search import native_search, get_dynamics_model, page_search
from common.utils.common import get_file_content
from knowledge.models import Knowledge
from maxkb.conf import PROJECT_DIR
from models_provider.base_model_provider import ModelTypeConst
from models_provider.models import Model
from system_manage.models import WorkspaceUserResourcePermission
from tools.models import Tool, ToolType

_PERM_WITH_ROLE = ["VIEW", "MANAGE", "ROLE"]
_PERM_DEFAULT = ["VIEW", "MANAGE"]
TOKEN_EXPR = F("chatrecord__message_tokens") + F("chatrecord__answer_tokens")


def hasPermission(auth, permission):
    if 'USER' in auth.role_list:
        return True
    if permission in auth.permission_list:
        return True
    return False


def has_extends_workspace_manage_permission(auth, permission, workspace_id):
    return hasPermission(auth, f"{permission}:/WORKSPACE/{workspace_id}:ROLE/WORKSPACE_MANAGE")


def has_user_permission(auth, permission, workspace_id):
    return hasPermission(auth, f"{permission}:/WORKSPACE/{workspace_id}")


def has_all_permission(auth, permission, workspace_id):
    return (has_user_permission(auth, permission, workspace_id)
            or has_extends_workspace_manage_permission(auth,
                                                       permission,
                                                       workspace_id)
            or hasPermission(auth,
                             permission)
            or RoleConstants.USER.name + f':/WORKSPACE/{workspace_id}' in auth.role_list
            or RoleConstants.WORKSPACE_MANAGE.name + f':/WORKSPACE/{workspace_id}' in auth.role_list)


def is_workspace_manage(auth, workspace_id):
    return RoleConstants.WORKSPACE_MANAGE.value.__str__() + ":/WORKSPACE/" + workspace_id in auth.role_list


def is_extends_workspace_manage(auth, workspace_id):
    return RoleConstants.EXTENDS_WORKSPACE_MANAGE.value.__str__() + ":/WORKSPACE/" + workspace_id in auth.role_list


def get_start_time(date_time):
    d = datetime.datetime.strptime(date_time, '%Y-%m-%d').date()
    naive = datetime.datetime.combine(d, datetime.time.min)
    return timezone.make_aware(naive, timezone.get_default_timezone())


def get_end_time(date_time):
    d = datetime.datetime.strptime(date_time, '%Y-%m-%d').date()
    naive = datetime.datetime.combine(d, datetime.time.max)
    return timezone.make_aware(naive, timezone.get_default_timezone())


class HomePageSerializer(serializers.Serializer):
    class ChatRecordAggregation(serializers.Serializer):
        workspace_id = serializers.CharField(required=False, label=_("Workspace ID"))
        user_id = serializers.UUIDField(required=True, label=_("User ID"))
        start_time = serializers.DateField(format='%Y-%m-%d', label=_("Start time"))
        end_time = serializers.DateField(format='%Y-%m-%d', label=_("End time"))

        def aggregation(self, auth, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            data = self.data
            user_id = data["user_id"]
            workspace_id = data.get("workspace_id")
            start_time = get_start_time(data.get('start_time'))
            end_time = get_end_time(data.get('end_time'))
            workspace_manage = is_workspace_manage(auth, workspace_id)
            extends_workspace_manage = is_extends_workspace_manage(auth, workspace_id)
            query = ChatRecord.objects.filter(
                create_time__gte=start_time,
                create_time__lte=end_time,
            )
            if workspace_manage:
                query = query.filter(
                    chat__application__workspace_id=workspace_id
                )
            elif extends_workspace_manage:
                if hasPermission(auth, f"APPLICATION:READ:/WORKSPACE/{workspace_id}"):
                    query = query.filter(
                        chat__application__workspace_id=workspace_id
                    )
                else:
                    return 0
            else:
                permission_list = (
                    ["VIEW", "MANAGE", "ROLE"]
                    if hasPermission(auth, "APPLICATION:READ")
                    else ["VIEW", "MANAGE"]
                )
                permission_subquery = (
                    WorkspaceUserResourcePermission.objects
                    .filter(
                        workspace_id=workspace_id,
                        user_id=user_id,
                        auth_target_type="APPLICATION",
                        permission_list__overlap=permission_list
                    ).exclude(target='default')
                    .annotate(
                        target_uuid=Cast(
                            "target",
                            output_field=UUIDField()
                        )
                    )
                    .values("target_uuid")
                )
                query = query.filter(
                    chat__application_id__in=permission_subquery
                )

            return query.aggregate(
                total_count=Count("id")
            )["total_count"]

    class TokensAggregation(serializers.Serializer):
        workspace_id = serializers.CharField(required=False, label=_("Workspace ID"))
        user_id = serializers.UUIDField(required=True, label=_("User ID"))
        start_time = serializers.DateField(format='%Y-%m-%d', label=_("Start time"))
        end_time = serializers.DateField(format='%Y-%m-%d', label=_("End time"))

        def aggregation(self, auth, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            data = self.data
            user_id = data["user_id"]
            workspace_id = data.get("workspace_id")
            start_time = get_start_time(data["start_time"])
            end_time = get_end_time(data["end_time"])
            workspace_manage = is_workspace_manage(auth, workspace_id)
            extends_workspace_manage = is_extends_workspace_manage(auth, workspace_id)
            query = ChatRecord.objects.filter(
                create_time__gte=start_time,
                create_time__lte=end_time,
            )
            if workspace_manage:
                query = query.filter(
                    chat__application__workspace_id=workspace_id
                )
            elif extends_workspace_manage and has_extends_workspace_manage_permission(auth, 'APPLICATION:READ',
                                                                                      workspace_id):
                query = query.filter(
                    chat__application__workspace_id=workspace_id
                )
            else:
                permission_list = (
                    ["VIEW", "MANAGE", "ROLE"]
                    if hasPermission(auth, "APPLICATION:READ")
                    else ["VIEW", "MANAGE"]
                )
                permission_subquery = (
                    WorkspaceUserResourcePermission.objects
                    .filter(
                        workspace_id=workspace_id,
                        user_id=user_id,
                        auth_target_type="APPLICATION",
                        permission_list__overlap=permission_list
                    ).exclude(target='default')
                    .annotate(
                        target_uuid=Cast(
                            "target",
                            output_field=UUIDField()
                        )
                    )
                    .values("target_uuid")
                )
                query = query.filter(
                    chat__application_id__in=permission_subquery
                )

            return query.aggregate(
                total_tokens=Coalesce(
                    Sum(
                        F("message_tokens") + F("answer_tokens"),
                        output_field=IntegerField()
                    ),
                    0
                )
            )["total_tokens"]

    class ApplicationUserTokenRanking(serializers.Serializer):
        workspace_id = serializers.CharField(required=False, label=_("Workspace ID"))
        user_id = serializers.UUIDField(required=True, label=_("User ID"))
        start_time = serializers.DateField(format='%Y-%m-%d', label=_("Start time"))
        name = serializers.CharField(required=False, allow_null=True, allow_blank=True, label=_("User Name"))
        end_time = serializers.DateField(format='%Y-%m-%d', label=_("End time"))

        def get_queryset(self, auth):
            workspace_id = self.data.get("workspace_id")
            user_id = self.data.get("user_id")
            start_time = get_start_time(self.data.get("start_time"))
            end_time = get_end_time(self.data.get("end_time"))
            name = self.data.get("name")

            # ---- 基础查询：不再按 Chat.create_time 过滤 ----
            base_queryset = (
                Chat.objects.filter(
                    is_deleted=False,
                    chat_user_id__isnull=False,
                )
                .exclude(chat_user_id="")
            )

            if name:
                base_queryset = base_queryset.filter(asker__username__contains=name)

            # ---- 权限过滤 ----
            base_queryset = self._apply_permission_filter(
                base_queryset, auth, workspace_id, user_id
            )

            # ---- 窗口函数：一次查询拿到每个用户最新的 asker ----
            asker_map = self._build_asker_map(base_queryset)

            # ---- 时间条件针对 ChatRecord ----
            record_time_filter = Q(
                chatrecord__create_time__gte=start_time,
                chatrecord__create_time__lte=end_time,
            )

            # ---- 聚合统计 ----
            queryset = (
                base_queryset
                .filter(record_time_filter)
                .values("chat_user_id", "chat_user_type")
                .annotate(
                    total_tokens=Coalesce(
                        Sum(TOKEN_EXPR, filter=record_time_filter),
                        Value(0),
                        output_field=BigIntegerField(),
                    ),
                    chat_record_count=Count(
                        "chatrecord__id",
                        filter=record_time_filter,
                        distinct=True,
                    ),
                )
                .order_by("-total_tokens")
            )
            return queryset, asker_map

        def ranking(self, auth, current_page, page_size, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            queryset, asker_map = self.get_queryset(auth)
            return page_search(
                current_page,
                page_size,
                queryset,
                lambda item: {
                    "chat_user_id": item["chat_user_id"],
                    "chat_user_type": item["chat_user_type"],
                    "asker": asker_map.get(
                        (item["chat_user_id"], item["chat_user_type"])
                    ),
                    "total_tokens": item["total_tokens"],
                    "chat_record_count": item["chat_record_count"],
                },
            )

        def export(self, auth, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            token_count = HomePageSerializer.TokensAggregation(data=self.data).aggregation(auth)
            queryset, asker_map = self.get_queryset(auth)
            workbook = openpyxl.Workbook(write_only=True)
            worksheet = workbook.create_sheet(title='Sheet1')
            headers = [gettext('ranking'),
                       gettext('User Name'),
                       gettext('Token consumption'),
                       gettext('proportion'),
                       gettext('number of questions'),
                       gettext('Average tokens per request'),
                       ]
            worksheet.append(headers)
            index = 0
            for item in queryset:
                index += 1
                user_info = asker_map.get(
                    (item["chat_user_id"], item["chat_user_type"])
                ) or {}
                username = user_info.get("username", "")
                total_tokens = item.get("total_tokens", 0)
                chat_record_count = item.get("chat_record_count", 0)
                row = [
                    index,
                    username,
                    total_tokens,
                    total_tokens / token_count if token_count else 0,
                    chat_record_count,
                    total_tokens / chat_record_count if chat_record_count else 0,
                ]
                worksheet.append(row)
            response = HttpResponse(content_type="application/vnd.ms-excel")
            response["Content-Disposition"] = f'attachment; filename="data.xlsx"'
            workbook.save(response)
            return response

        def _apply_permission_filter(self, queryset, auth, workspace_id, user_id):
            """根据用户角色过滤可见的应用范围"""
            if is_workspace_manage(auth, workspace_id):
                return queryset.filter(application__workspace_id=workspace_id)
            elif is_extends_workspace_manage(auth, workspace_id):
                if hasPermission(auth, f"APPLICATION:READ:/WORKSPACE/{workspace_id}"):
                    return queryset.filter(application__workspace_id=workspace_id)
            if not has_all_permission(auth, 'APPLICATION:READ', workspace_id):
                return queryset.none()

            permission_list = (
                _PERM_WITH_ROLE
                if hasPermission(auth, "APPLICATION:READ")
                else _PERM_DEFAULT
            )

            allowed_app_ids = (
                QuerySet(WorkspaceUserResourcePermission)
                .filter(
                    workspace_id=workspace_id,
                    user_id=user_id,
                    auth_target_type="APPLICATION",
                    permission_list__overlap=permission_list,
                ).exclude(target='default')
                .annotate(target_uuid=Cast("target", output_field=UUIDField()))
                .values_list("target_uuid", flat=True)
            )

            return queryset.filter(application_id__in=allowed_app_ids)

        @staticmethod
        def _build_asker_map(base_queryset):
            """
            用窗口函数一次查询拿到每个 (chat_user_id, chat_user_type) 最新的 asker，
            替代原来每行一次的 Subquery。
            """
            latest_rows = (
                base_queryset
                .annotate(
                    _rn=Window(
                        expression=RowNumber(),
                        partition_by=[F("chat_user_id"), F("chat_user_type")],
                        order_by=F("create_time").desc(),
                    )
                )
                .filter(_rn=1)
                .values("chat_user_id", "chat_user_type", "asker")
            )

            return {
                (row["chat_user_id"], row["chat_user_type"]): row["asker"]
                for row in latest_rows
            }

    class ApplicationQuestionRanking(serializers.Serializer):
        workspace_id = serializers.CharField(required=False, label=_('Workspace ID'))
        user_id = serializers.UUIDField(required=True, label=_("User ID"))
        name = serializers.CharField(required=False, allow_null=True, allow_blank=True, label=_("Application Name"))
        start_time = serializers.DateField(format='%Y-%m-%d', label=_("Start time"))
        end_time = serializers.DateField(format='%Y-%m-%d', label=_("End time"))

        def get_queryset(self, auth):
            workspace_id = self.data.get("workspace_id")
            user_id = self.data.get("user_id")
            name = self.data.get("name")
            start_time = get_start_time(self.data.get("start_time"))
            end_time = get_end_time(self.data.get("end_time"))
            workspace_manage = is_workspace_manage(auth, workspace_id)
            queryset = QuerySet(Application)
            is_resource_filter = True
            if name:
                queryset = queryset.filter(name__contains=name)
                is_resource_filter = False
            if workspace_manage:
                queryset = queryset.filter(workspace_id=workspace_id)
            elif is_extends_workspace_manage(auth, workspace_id):
                if has_extends_workspace_manage_permission(auth, "APPLICATION:READ", workspace_id):
                    queryset = queryset.filter(workspace_id=workspace_id)
                    is_resource_filter = False
            if not has_all_permission(auth, 'APPLICATION:READ', workspace_id):
                queryset = queryset.none()
                is_resource_filter = False
            if is_resource_filter:
                permission_list = (
                    ["VIEW", "MANAGE", "ROLE"]
                    if hasPermission(auth, "APPLICATION:READ")
                    else ["VIEW", "MANAGE"]
                )

                queryset = queryset.filter(
                    id__in=QuerySet(WorkspaceUserResourcePermission)
                    .filter(
                        workspace_id=workspace_id,
                        user_id=user_id,
                        auth_target_type="APPLICATION",
                        permission_list__overlap=permission_list,
                    ).exclude(target='default')
                    .annotate(
                        target_uuid=Cast("target", output_field=UUIDField())
                    )
                    .values_list("target_uuid", flat=True)
                )

            record_time_filter = (
                    Q(chat__is_deleted=False)
                    & Q(chat__chatrecord__create_time__gte=start_time)
                    & Q(chat__chatrecord__create_time__lte=end_time)
            )
            return queryset.annotate(
                # 问题数（按 ChatRecord 条数统计）
                chat_record_count_total=Coalesce(
                    Count(
                        "chat__chatrecord__id",
                        filter=record_time_filter,
                    ),
                    Value(0),
                    output_field=BigIntegerField(),
                ),

                # 对话用户数量，按 chat_user_id 去重
                chat_user_count=Count(
                    "chat__chat_user_id",
                    filter=(
                            record_time_filter
                            & Q(chat__chat_user_id__isnull=False)
                            & ~Q(chat__chat_user_id="")
                    ),
                    distinct=True,
                ),
            ).order_by(
                "-chat_record_count_total"
            )

        def ranking(self, auth, current_page, page_size, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            queryset = self.get_queryset(auth)
            return page_search(
                current_page,
                page_size,
                queryset,
                lambda a: {
                    "id": a.id,
                    "name": a.name,
                    "chat_record_count": a.chat_record_count_total,
                    "chat_user_count": a.chat_user_count,
                },
            )

        def export(self, auth, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            chat_record_number = HomePageSerializer.ChatRecordAggregation(data=self.data).aggregation(auth)
            queryset = self.get_queryset(auth)
            workbook = openpyxl.Workbook(write_only=True)
            worksheet = workbook.create_sheet(title='Sheet1')
            headers = [gettext('ranking'),
                       gettext('Application Name'),
                       gettext('number of questions'),
                       gettext('proportion'),
                       gettext('active users'),
                       gettext('Average Number of Conversation Turns per Person')
                       ]
            worksheet.append(headers)
            index = 0
            for item in queryset:
                index += 1
                row = [
                    index,
                    item.name,
                    item.chat_record_count_total,
                    item.chat_record_count_total / chat_record_number if chat_record_number != 0 else 0,
                    item.chat_user_count,
                    item.chat_user_count / item.chat_record_count_total if item.chat_record_count_total != 0 else 0
                ]
                worksheet.append(row)
            response = HttpResponse(content_type="application/vnd.ms-excel")
            response["Content-Disposition"] = f'attachment; filename="data.xlsx"'
            workbook.save(response)
            return response

    class ApplicationTokensRanking(serializers.Serializer):
        workspace_id = serializers.CharField(required=False, label=_('Workspace ID'))
        user_id = serializers.UUIDField(required=True, label=_("User ID"))
        name = serializers.CharField(required=False, allow_null=True, allow_blank=True, label=_("Application Name"))
        start_time = serializers.DateField(format='%Y-%m-%d', label=_("Start time"))
        end_time = serializers.DateField(format='%Y-%m-%d', label=_("End time"))

        def get_queryset(self, auth):
            start_time = get_start_time(self.data.get('start_time'))
            end_time = get_end_time(self.data.get('end_time'))
            name = self.data.get("name")
            workspace_id = self.data.get("workspace_id")
            user_id = self.data.get("user_id")

            token_expr = ExpressionWrapper(
                F("chat__chatrecord__message_tokens") + F("chat__chatrecord__answer_tokens"),
                output_field=BigIntegerField()
            )

            # 时间条件针对 ChatRecord
            record_time_filter = (
                    Q(chat__is_deleted=False)
                    & Q(chat__chatrecord__create_time__gte=start_time)
                    & Q(chat__chatrecord__create_time__lte=end_time)
            )
            is_resource_filter = True
            workspace_manage = is_workspace_manage(auth, workspace_id)
            queryset = QuerySet(Application)
            if name:
                queryset = queryset.filter(name__contains=name)
            if workspace_manage:
                queryset = queryset.filter(workspace_id=workspace_id)
                is_resource_filter = False
            elif is_extends_workspace_manage(auth, workspace_id):
                if has_extends_workspace_manage_permission(
                        auth,
                        "APPLICATION:READ", workspace_id
                ):
                    queryset = queryset.filter(workspace_id=workspace_id)
                    is_resource_filter = False
            if not has_all_permission(auth, 'APPLICATION:READ', workspace_id):
                queryset = queryset.none()
                is_resource_filter = False

            if is_resource_filter:
                permission_list = ["VIEW", "MANAGE", "ROLE"] if hasPermission(
                    auth,
                    "APPLICATION:READ"
                ) else ["VIEW", "MANAGE"]

                queryset = queryset.filter(
                    id__in=QuerySet(WorkspaceUserResourcePermission)
                    .filter(
                        workspace_id=workspace_id,
                        user_id=user_id,
                        auth_target_type="APPLICATION",
                        permission_list__overlap=permission_list
                    ).exclude(target='default')
                    .annotate(target_uuid=Cast("target", output_field=UUIDField()))
                    .values_list("target_uuid", flat=True)
                )

            return queryset.annotate(
                total_tokens=Coalesce(
                    Sum(
                        token_expr,
                        filter=record_time_filter
                    ),
                    Value(0),
                    output_field=BigIntegerField()
                ),
                chat_record_count_total=Count(
                    "chat__chatrecord__id",
                    filter=record_time_filter,
                    output_field=IntegerField()
                ),
                chat_user_count=Count(
                    "chat__chat_user_id",
                    filter=(
                            record_time_filter
                            & Q(chat__chat_user_id__isnull=False)
                            & ~Q(chat__chat_user_id="")
                    ),
                    distinct=True,
                ),
            ).order_by("-total_tokens")

        def ranking(self, auth, current_page, page_size, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            queryset = self.get_queryset(auth)
            return page_search(
                current_page,
                page_size,
                queryset,
                lambda a: {
                    "id": a.id,
                    "name": a.name,
                    "total_tokens": a.total_tokens,
                    "chat_record_count": a.chat_record_count_total,
                    "chat_user_count": a.chat_user_count
                }
            )

        def export(self, auth, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            tokens_total = HomePageSerializer.TokensAggregation(data=self.data).aggregation(auth)
            queryset = self.get_queryset(auth)
            workbook = openpyxl.Workbook(write_only=True)
            worksheet = workbook.create_sheet(title='Sheet1')
            headers = [gettext('ranking'),
                       gettext('Application Name'),
                       gettext('Token consumption'),
                       gettext('proportion'),
                       gettext('number of questions'),
                       gettext('active users'),
                       gettext('Average tokens per request'),
                       ]
            worksheet.append(headers)
            index = 0
            for item in queryset:
                index += 1
                total_tokens = item.total_tokens
                chat_record_count_total = item.chat_record_count_total
                row = [
                    index,
                    item.name,
                    total_tokens,
                    total_tokens / tokens_total if tokens_total else 0,
                    item.chat_user_count,
                    chat_record_count_total,
                    total_tokens / chat_record_count_total if chat_record_count_total else 0,
                ]
                worksheet.append(row)
            response = HttpResponse(content_type="application/vnd.ms-excel")
            response["Content-Disposition"] = f'attachment; filename="data.xlsx"'
            workbook.save(response)
            return response

    class ApplicationMonitoring(serializers.Serializer):
        workspace_id = serializers.CharField(required=False, label=_('Workspace ID'))
        user_id = serializers.UUIDField(required=True, label=_("User ID"))
        application_id = serializers.UUIDField(required=False, allow_null=True, label=_("Application ID"))
        start_time = serializers.DateField(format='%Y-%m-%d', label=_("Start time"))
        end_time = serializers.DateField(format='%Y-%m-%d', label=_("End time"))

        def get_customer_count_trend(self, application_queryset, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            start_time = get_start_time(self.data.get("start_time"))
            end_time = get_end_time(self.data.get("end_time"))
            query_set = QuerySet(ApplicationChatUserStats).filter(
                create_time__gte=start_time,
                create_time__lte=end_time)
            application_id = self.data.get('application_id')
            if application_id:
                query_set = query_set.filter(application_id=application_id)
            else:
                query_set = query_set.filter(application_id__in=application_queryset)
            return native_search(
                {'default_sql': query_set},
                select_string=get_file_content(
                    os.path.join(PROJECT_DIR, "apps", "application", 'sql', 'customer_count_trend.sql')))

        def get_chat_record_aggregate_trend(self, auth, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            user_id = self.data.get("user_id")
            workspace_id = self.data.get("workspace_id")
            start_time = get_start_time(self.data.get("start_time"))
            end_time = get_end_time(self.data.get("end_time"))
            application_id = self.data.get('application_id')
            applicationSerializer = HomePageSerializer.Application(
                data={"user_id": user_id, 'workspace_id': workspace_id})
            applicationSerializer.is_valid(raise_exception=True)
            application_query_set = applicationSerializer.get_aggregation_query_set(
                auth)
            chat_record_aggregate_trend = native_search(
                {'default_sql': QuerySet(model=get_dynamics_model(
                    {'application_chat.application_id': models.UUIDField(),
                     'application_chat_record.create_time': models.DateTimeField()})).filter(
                    **{**({'application_chat.application_id': application_id} if application_id else {
                        'application_chat.application_id__in': application_query_set}),
                       'application_chat_record.create_time__gte': start_time,
                       'application_chat_record.create_time__lte': end_time}
                )},
                select_string=get_file_content(
                    os.path.join(PROJECT_DIR, "apps", "application", 'sql', 'chat_record_count_trend.sql')))
            customer_count_trend = self.get_customer_count_trend(application_query_set, with_valid=False)
            return self.merge_customer_chat_record(chat_record_aggregate_trend, customer_count_trend)

        def merge_customer_chat_record(self, chat_record_aggregate_trend: List[Dict], customer_count_trend: List[Dict]):

            return [{**self.find(chat_record_aggregate_trend, lambda c: c.get('day').strftime('%Y-%m-%d') == day,
                                 {'star_num': 0, 'trample_num': 0, 'tokens_num': 0, 'chat_record_count': 0,
                                  'customer_num': 0,
                                  'day': day}),
                     **self.find(customer_count_trend, lambda c: c.get('day').strftime('%Y-%m-%d') == day,
                                 {'customer_added_count': 0})}
                    for
                    day in
                    self.get_days_between_dates(self.data.get('start_time'), self.data.get('end_time'))]

        @staticmethod
        def find(source_list, condition, default):
            value_list = [row for row in source_list if condition(row)]
            if len(value_list) > 0:
                return value_list[0]
            return default

        @staticmethod
        def get_days_between_dates(start_date, end_date):
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
            days = []
            current_date = start_date
            while current_date <= end_date:
                days.append(current_date.strftime('%Y-%m-%d'))
                current_date += datetime.timedelta(days=1)
            return days

    class Application(serializers.Serializer):
        workspace_id = serializers.CharField(required=False, label=_('Workspace ID'))
        user_id = serializers.UUIDField(required=True, label=_("User ID"))

        def get_aggregation_query_set(self, auth):
            workspace_id = self.data.get("workspace_id")
            user_id = self.data.get("user_id")
            workspace_manage = is_workspace_manage(auth, workspace_id)
            if workspace_manage:
                return QuerySet(Application).filter(workspace_id=workspace_id)
            if is_extends_workspace_manage(auth, workspace_id):
                if has_extends_workspace_manage_permission(auth, "APPLICATION:READ", workspace_id):
                    return QuerySet(Application).filter(workspace_id=workspace_id)
            if not has_all_permission(auth, 'APPLICATION:READ', workspace_id):
                return QuerySet(Application).none()
            permission_list = ["VIEW", "MANAGE", "ROLE"] if hasPermission(auth, "APPLICATION:READ") else ['VIEW',
                                                                                                          'MANAGE']
            return QuerySet(Application).filter(
                id__in=QuerySet(WorkspaceUserResourcePermission)
                .filter(workspace_id=workspace_id,
                        user_id=user_id,
                        auth_target_type="APPLICATION",
                        permission_list__overlap=permission_list
                        ).exclude(target='default').annotate(target_uuid=Cast("target", output_field=UUIDField()))
                .values_list("target_uuid", flat=True))

        def aggregation(self, auth, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            query_set = self.get_aggregation_query_set(auth)
            result = query_set.aggregate(
                total=Count("id"),
                publish_count=Count("id", filter=Q(is_publish=True)),
                un_publish_count=Count("id", filter=Q(is_publish=False)),
            )
            return {
                "total": result["total"],
                "publish_count": result["publish_count"],
                "un_publish_count": result["un_publish_count"],
            }

    class Knowledge(serializers.Serializer):
        workspace_id = serializers.CharField(required=False, label=_('Workspace ID'))
        user_id = serializers.UUIDField(required=True, label=_("User ID"))

        def get_aggregation_query_set(self, auth):
            workspace_id = self.data.get("workspace_id")
            user_id = self.data.get("user_id")
            if is_workspace_manage(auth, workspace_id):
                return QuerySet(Knowledge).filter(workspace_id=workspace_id)
            if is_extends_workspace_manage(auth, workspace_id):
                if has_extends_workspace_manage_permission(auth, "KNOWLEDGE:READ", workspace_id):
                    return QuerySet(Knowledge).filter(workspace_id=workspace_id)
            if not has_all_permission(auth, 'KNOWLEDGE:READ', workspace_id):
                return QuerySet(Knowledge).none()
            permission_list = ["VIEW", "MANAGE", "ROLE"] if hasPermission(auth, "KNOWLEDGE:READ") else ['VIEW',
                                                                                                        'MANAGE']
            return QuerySet(Knowledge).filter(
                id__in=QuerySet(WorkspaceUserResourcePermission).filter(workspace_id=workspace_id,
                                                                        user_id=user_id,
                                                                        auth_target_type="KNOWLEDGE",
                                                                        permission_list__overlap=permission_list
                                                                        ).exclude(target='default').annotate(
                    target_uuid=Cast("target", output_field=UUIDField()))
                .values_list("target_uuid", flat=True))

        def aggregation(self, auth, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            query_set = self.get_aggregation_query_set(auth)
            result = query_set.aggregate(
                total=Count("id", distinct=True),
                document_count=Count(
                    "document",
                    distinct=True,
                ),
                failure_count=Count(
                    "document",
                    filter=Q(
                        document__status__contains="3",
                    ),
                    distinct=True,
                ),
            )
            return {
                "total": result["total"] or 0,
                "document_count": result["document_count"] or 0,
                "failure_count": result["failure_count"] or 0,
            }

    class Tool(serializers.Serializer):
        workspace_id = serializers.CharField(required=False, label=_('Workspace ID'))
        user_id = serializers.UUIDField(required=True, label=_("User ID"))

        def get_aggregation_query_set(self, auth):
            workspace_id = self.data.get("workspace_id")
            user_id = self.data.get("user_id")
            if is_workspace_manage(auth, workspace_id):
                return QuerySet(Tool).filter(workspace_id=workspace_id)
            if is_extends_workspace_manage(auth, workspace_id):
                if has_extends_workspace_manage_permission(auth, "TOOL:READ", workspace_id):
                    return QuerySet(Tool).filter(workspace_id=workspace_id)
            if not has_all_permission(auth, 'TOOL:READ', workspace_id):
                return QuerySet(Tool).none()
            permission_list = ["VIEW", "MANAGE", "ROLE"] if hasPermission(auth, "TOOL:READ") else ['VIEW',
                                                                                                   'MANAGE']
            return QuerySet(Tool).filter(
                id__in=QuerySet(WorkspaceUserResourcePermission).filter(workspace_id=workspace_id,
                                                                        user_id=user_id,
                                                                        auth_target_type="TOOL",
                                                                        permission_list__overlap=permission_list
                                                                        )
                .exclude(target='default').annotate(
                    target_uuid=Cast("target", output_field=UUIDField()))
                .values_list("target_uuid", flat=True))

        def aggregation(self, auth, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            query_set = self.get_aggregation_query_set(auth)
            result = query_set.aggregate(
                total=Count("id"),
                custom_count=Count("id", filter=Q(tool_type=ToolType.CUSTOM)),
                skill_count=Count("id", filter=Q(tool_type=ToolType.SKILL)),
                mcp_count=Count("id", filter=Q(tool_type=ToolType.MCP)),
                workflow_count=Count("id", filter=Q(tool_type=ToolType.WORKFLOW)),
                data_source_count=Count("id", filter=Q(tool_type=ToolType.DATA_SOURCE)),
            )
            return {
                "total": result["total"] or 0,
                "custom_count": result["custom_count"] or 0,
                "skill_count": result["skill_count"] or 0,
                "mcp_count": result["mcp_count"] or 0,
                "workflow_count": result["workflow_count"] or 0,
                "data_source_count": result["data_source_count"] or 0,
            }

    class Model(serializers.Serializer):
        workspace_id = serializers.CharField(required=False, label=_('Workspace ID'))
        user_id = serializers.UUIDField(required=True, label=_("User ID"))

        def get_aggregation_query_set(self, auth):
            workspace_id = self.data.get("workspace_id")
            user_id = self.data.get("user_id")
            if is_workspace_manage(auth, workspace_id):
                return QuerySet(Model).filter(workspace_id=workspace_id)
            if is_extends_workspace_manage(auth, workspace_id):
                if has_extends_workspace_manage_permission(auth, "MODEL:READ", workspace_id):
                    return QuerySet(Model).filter(workspace_id=workspace_id)
            if not has_all_permission(auth, 'MODEL:READ', workspace_id):
                return QuerySet(Model).none()
            permission_list = ["VIEW", "MANAGE", "ROLE"] if hasPermission(auth, "MODEL:READ") else ['VIEW',
                                                                                                    'MANAGE']
            return QuerySet(Model).filter(
                id__in=QuerySet(WorkspaceUserResourcePermission).filter(workspace_id=workspace_id,
                                                                        user_id=user_id,
                                                                        auth_target_type="MODEL",
                                                                        permission_list__overlap=permission_list
                                                                        ).exclude(target='default').annotate(
                    target_uuid=Cast("target", output_field=UUIDField()))
                .values_list("target_uuid", flat=True))

        def aggregation(self, auth, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            query_set = self.get_aggregation_query_set(auth)
            result = query_set.aggregate(
                total=Count("id"),
                embedding_count=Count("id", filter=Q(model_type=ModelTypeConst.EMBEDDING.name)),
                llm_count=Count("id", filter=Q(model_type=ModelTypeConst.LLM.name)),
            )
            total = result["total"] or 0
            embedding_count = result["embedding_count"] or 0
            llm_count = result["llm_count"] or 0
            return {
                "total": total,
                "embedding_count": embedding_count,
                "llm_count": llm_count
            }
